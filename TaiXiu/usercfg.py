from openpyxl import load_workbook

def HAdmin():
    options = input('p/c/w/m/per/q | (Play/CreateAcc/Wr/Money/Permission/quit): ')
    while ( options != 'p' and options != 'c' and options != 'w' and options != 'm' and options != 'per' and options != 'q'): 
        options = input('p/c/w/m/per/q (Play/CreateAcc/Wr/Money/Permission/quit) | Please input again: ')

    if(options == 'q'):
        return False
    
    if(options == 'p'):
        import gp
        gp.game(2)
        return

    if(options == 'c'): 
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        count = ws['E2'].value + 1
        ws['A' + str(count)] = count - 1


        def UserNC():
            options = input('C/Q (Continue / Quit): ')
            if options == 'q' or options == 'Q':
                import sys
                sys.exit()
            username = input('Enter your username: ')
            for i in range(2, count+1, 1):
                if username == ws['B' + str(i)].value: 
                    print("This username has been used!")
                    return False
            ws['B' + str(count)] = username
            wb.save('Data.xlsx')
            return True
        while not UserNC():
            if UserNC(): break


        ws['C' + str(count)] = input('Input password: ')
        bal = input('Bal: ')
        while not bal.isdigit() or int(bal) < 0:
            bal = input('Input again: ')
        ws['D' + str(count)] = bal
        per = input('1/2 | Admin/User: ')
        while not per.isdigit() or 1 < int(per) > 2: 
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                return
            per = input('Input again: ')
        ws['F' + str(count)] = int(per)
        print('The Account has been created succesfully')
        ws['E2'] = count
        wb.save('Data.xlsx')
        return
    
    if options == 'w': 
        wr = input('(0/1/6) | Config wr | Normal/Xiu/Tai: ')
        while not wr.isdigit() or (int(wr) != 0 and int(wr) != 1 and int(wr) != 6): 
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                return
            wr = input('(0/1/6) | Config wr | Normal/Xiu/Tai | Input again: ')
        print("Config successfully")
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        ws['H1'] = wr
        wb.save('Data.xlsx')
        return
    
    if options == 'm':
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        count = ws['E2'].value + 1

        def UserNC():
            options = input('C/Q (Continue / Quit): ')
            if options == 'q' or options == 'Q':
                import sys
                sys.exit()

            username = input('Enter your username: ')
            for i in range(2, count+1, 1):
                wb = load_workbook('Data.xlsx')
                ws = wb.active
                if username == ws['B' + str(i)].value:
                    wb = load_workbook('Data.xlsx')
                    ws = wb.active 
                    bal = input('Bal: ')
                    while not bal[1:len(bal)].isdigit():
                        options = input('input C/Q (continue/quit): ')
                        if(options == 'q' or options == 'Q'):
                            return  False
                        bal = input('Enter again | Bal: ')
                        
                    if int(bal) > 0: 
                        wb = load_workbook('Data.xlsx')
                        ws = wb.active
                        NewMon = (int(ws['D' + str(i)].value) + int(bal))
                        ws['D' + str(i)] = NewMon
                        wb.save('Data.xlsx')
                        print('Add successfully!')
                        return  True
                    else:
                        wb = load_workbook('Data.xlsx')
                        ws = wb.active
                        NewMon = (int(ws['D' + str(i)].value) + int(bal))
                        ws['D' + str(i)] = NewMon
                        wb.save('Data.xlsx')
                        print('Withdraw successfully!')
                        return True
            return False
        
        while not UserNC():
            if UserNC(): break
        
    
    if options == 'per':
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        count = ws['E2'].value + 1 
        for i in range (3, count, 1):
            print('%s | %s'%(ws['B' + str(i)].value, ws['F' + str(i)].value ))
        wb.save('Data.xlsx')  
        
        username = input('Username: ')
        for i in range (3, count, 1):
            if username == ws['B' + str(i)].value:
                per = input('Per: ')
                while not per.isdigit() or (int(per) != 0 and int(per) != 1 and int(per) != 2):
                    options = input('input C/Q (continue/quit): ')
                    if(options == 'q' or options == 'Q'):
                        return  
                    per = input('Input again: ')
                for i in range(2, count, 1):
                    wb = load_workbook('Data.xlsx')
                    ws = wb.active
                    if username == ws['B' + str(i)].value:
                        ws['F' + str(i)] = int(per)
                        print('Change successfully')
                        wb.save('Data.xlsx')
                        return
            else: print('Config failed')
            return


def Admin(id):
    options = input('p/c/w/m/q | (Play/CreateAcc/Wr/Money/quit): ')
    while ( options != 'p' and options != 'c' and options != 'w' and options != 'm' and options != 'q'): 
        options = input('p/c/w/m/q (Play/CreateAcc/Wr/Money/quit) | Please input again: ')

    if(options == 'q'):
        return False
    
    if(options == 'p'):
        import gp
        gp.game(id)
        return

    if(options == 'c'): 
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        count = ws['E2'].value + 1
        ws['A' + str(count)] = count - 1


        def UserNC():
            options = input('C/Q (Continue / Quit): ')
            if options == 'q' or options == 'Q':
                import sys
                sys.exit()
            username = input('Enter your username: ')
            for i in range(2, count+1, 1):
                if username == ws['B' + str(i)].value: 
                    print("This username has been used!")
                    return False
            ws['B' + str(count)] = username
            wb.save('Data.xlsx')
            return True
        while not UserNC():
            if UserNC(): break


        ws['C' + str(count)] = input('Input password: ')
        bal = input('Bal: ')
        while not bal.isdigit() or int(bal) < 0:
            bal = input('Input again: ')
        ws['D' + str(count)] = bal
        per = input('1/2 | Admin/User: ')
        while not per.isdigit() or 1 < int(per) > 2: 
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                return
            per = input('Input again: ')
        ws['F' + str(count)] = int(per)
        print('The Account has been created succesfully')
        ws['E2'] = count
        wb.save('Data.xlsx')
        return
    
    if options == 'w': 
        wr = input('(0/1/6) | Config wr | Normal/Xiu/Tai: ')
        while not wr.isdigit() or (int(wr) != 0 and int(wr) != 1 and int(wr) != 6): 
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                return
            wr = input('(0/1/6) | Config wr | Normal/Xiu/Tai | Input again: ')
        print("Config successfully")
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        ws['H1'] = wr
        wb.save('Data.xlsx')
        return
    
    if options == 'm':
        wb = load_workbook('Data.xlsx')
        ws = wb.active
        count = ws['E2'].value + 1

        def UserNC():
            options = input('C/Q (Continue / Quit): ')
            if options == 'q' or options == 'Q':
                import sys
                sys.exit()

            username = input('Enter your username: ')
            for i in range(2, count+1, 1):
                wb = load_workbook('Data.xlsx')
                ws = wb.active
                if username == ws['B' + str(i)].value: 
                    bal = input('Bal: ')
                    while not bal[1:len(bal)].isdigit():
                        options = input('input C/Q (continue/quit): ')
                        if(options == 'q' or options == 'Q'):
                            return  False
                        bal = input('Enter again | Bal: ')

                    if int(bal) > 0: 
                        wb = load_workbook('Data.xlsx')
                        ws = wb.active
                        NewMon = (int(ws['D' + str(i)].value) + int(bal))
                        ws['D' + str(i)] = NewMon
                        wb.save('Data.xlsx')
                        print('Add successfully!')
                        return  True
                    else:
                        wb = load_workbook('Data.xlsx')
                        ws = wb.active
                        NewMon = (int(ws['D' + str(i)].value) + int(bal))
                        ws['D' + str(i)] = NewMon
                        wb.save('Data.xlsx')
                        print('Withdraw successfully!')
                        return True
            return False
        
        while not UserNC():
            if UserNC(): break
                
        

def user(id):
    options = input('p/m/q | (Play/Money/quit): ')
    while ( options != 'p' and options != 'm' and options != 'q'): 
        options = input('p/m/q (Play/Money) | Please input again: ')

    if(options == 'q'):
        return False
    
    if(options == 'p'):
        import gp
        gp.game(id)
        return
    
    if options == 'm':
        def usernc(): 
            wb = load_workbook('Data.xlsx')
            ws = wb.active

            mode = input('w/d | Withdraw/Deposit: ')
            while (mode != 'w' and mode != 'd'):
                mode = input('w/d | Withdraw/Deposit | Invalid value input again: ')
            
            bal = input('Bal: ')
            while not bal[1:len(bal)].isdigit():
                options = input('input C/Q (continue/quit): ')
                if(options == 'q' or options == 'Q'):
                    return  False
                bal = input('Enter again | Bal: ')
            
            if int(bal) > int(ws['D' +str(id)].value) and mode == 'w':
                print('Withdraw failed')
                return False
                
            if int(bal) > 0 and mode == 'd': 
                NewMon = (int(ws['D' + str(id)].value) + int(bal))
                wb = load_workbook('Data.xlsx')
                ws = wb.active
                ws['D' + str(id)] = NewMon
                wb.save('Data.xlsx')
                print('Add successfully!')
                return True
            elif mode == 'w' and int(bal) <= 0:
                NewMon = (int(ws['D' + str(id)].value) + int(bal))
                wb = load_workbook('Data.xlsx')
                ws = wb.active
                ws['D' + str(id)] = NewMon
                wb.save('Data.xlsx')
                print('Withdraw successfully!')
                return True
            return False
        while not usernc():
            if usernc(): break