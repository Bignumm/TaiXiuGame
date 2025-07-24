
from openpyxl import load_workbook

def login():
    LogedCheck = False
    while(LogedCheck == False):
        RegoLog = input("Reg/Log(Press R for Register / L for Login!): ")


        if(RegoLog == 'R' or RegoLog == 'r'):

            wb = load_workbook('Data.xlsx')
            ws = wb.active
            count = ws['E2'].value + 1


            def UserNC():
                options = input('C/Q (Continue / Quit): ')
                if options == 'q' or options == 'Q':
                    import sys
                    sys.exit()
                username = input('Enter your username: ')
                for i in range(2, count+1, 1):
                    if username == ws['B' + str(i)].value: 
                            print("This username has been used!")
                            return False
                ws['B' + str(count)] = username
                return True
            while not UserNC():
                if UserNC(): break

            ws['A' + str(count)] = count - 1
            ws['C' + str(count)] = input('Enter your password: ')


            bal = input('Add money: ')
            while(not bal.isdigit()):
                bal = input('Please enter again: ')
            ws['D' + str(count)] = bal 
            
            ws['E2'] = count
            
            ws['F' + str(count)] = 2

            wb.save('Data.xlsx')

            return True, count, ws['F' + str(count)].value
            


        elif(RegoLog == 'L' or RegoLog == 'l'):

            while(LogedCheck == False):
                wb = load_workbook('Data.xlsx')
                ws = wb.active
                username = input('Enter your username: ')
                password = input('Enter your password: ')
                count = ws['E2'].value + 1

                for i in range(2, count + 1, 1):
                    if username == ws['B' + str(i)].value and password == ws['C' + str(i)].value: 
                            print('Login successful!')
                            return True, i, ws['F' + str(i)].value
                wb.save('Data.xlsx')
                return False, None, False