
import Initset
import os.path


from openpyxl import load_workbook


if not os.path.exists('Data.xlsx'):
    Initset.Setup()

import login
import usercfg


status = True
while status:
    log, id, per = login.login()

    if not log:
        print('Account did not find!')
    
    while log:
        if per == 0:
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            usercfg.HAdmin()
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                log = False
            import os
            os.system('cls')

        elif per == 1:
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            usercfg.Admin(id)
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                log = False
            import os
            os.system('cls')

        elif per == 2:
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            usercfg.user(id)
            wb = load_workbook('Data.xlsx')
            ws = wb.active
            print('Account: %s\nBalance: %s' %(ws['B'+ str(id)].value, ws['D' + str(id)].value))
            wb.save('Data.xlsx')
            options = input('input C/Q (continue/quit): ')
            if(options == 'q' or options == 'Q'):
                log = False
            import os
            os.system('cls')

    options = input('input C/Q (continue/quit): ')
    if(options == 'q' or options == 'Q'):
        status = False