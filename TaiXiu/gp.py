import random
from openpyxl import load_workbook

def Bet(id):
    wb = load_workbook('Data.xlsx')
    ws = wb.active
    bet = int(input('1/0 | Tai/Xiu: '))
    while (bet != 1 and bet != 0): 
        bet = int(input('Invalid value | 1/0 | Tai/Xiu: '))
    mon = input('Input bet money: ')
    while not mon.isdigit() or int(mon) > int(ws['D' + str(id)].value): 
        options = input('input C/Q (continue/quit): ')
        if(options == 'q' or options == 'Q'):
            return 0, 0, 0
        mon = input('Invalid value | Input again: ')
    level = int(input('X? | max/min | 10/1: '))
    while(1 < level > 10):
        level = input('X? | max/min | 10/1: ')
    return bet, mon, level

def dice(): 
    wb = load_workbook('Data.xlsx')
    ws = wb.active
    
    if int(ws['H1'].value) == 1:
        a = random.randint(1, 4)
        b = random.randint(1, 3)
        c = random.randint(1, 3)
        res = a + b + c
        return res, a, b, c
    
    if int(ws['H1'].value) == 6:
        a = random.randint(4, 6)
        b = random.randint(3, 6)
        c = random.randint(4, 6)
        res = a + b + c
        return res, a, b, c

    if int(ws['H1'].value) == 0:
        a = random.randint(1, 6)
        b = random.randint(1, 6)
        c = random.randint(1, 6)
        res = a + b + c
        return res, a, b, c
    
def game(id):
    wb = load_workbook('Data.xlsx')
    ws = wb.active
    bet, mon, level = Bet(id)
    if(bet == 0 and mon == 0 and level == 0):
        return
    mon = int(mon)
    res, a, b, c = dice()
    if res > 10 and bet == 1:
        NewMon = (int(ws['D' + str(id)].value) + mon*level)
        print('%s / %s / %s / | Tai! | Thang | So du hien tai la: ' %(a, b, c) + str(NewMon))
        ws['D' + str(id)] = NewMon 
        print(NewMon)
        wb.save('Data.xlsx')
        return
    
    elif res > 10 and bet == 0:
        NewMon = (int(ws['D' + str(id)].value) - mon*level)
        print('%s / %s / %s / | Tai! | Thua | So du hien tai la: ' %(a, b, c) + str(NewMon))
        ws['D' + str(id)] = NewMon  
        print(NewMon)
        wb.save('Data.xlsx') 
        return
    
    elif res < 11 and bet == 1:
        NewMon = (int(ws['D' + str(id)].value) - mon*level)
        print('%s / %s / %s / | Xiu! | Thua | So du hien tai la: ' %(a, b, c) + str(NewMon))
        ws['D' + str(id)] = NewMon  
        print(NewMon)
        print(ws['D' + str(id)].value)
        wb.save('Data.xlsx') 
        return
    
    elif res < 11 and bet == 0:
        NewMon = (int(ws['D' + str(id)].value) + mon*level)
        print('%s / %s / %s / | Xiu! | Thang | So du hien tai la: ' %(a, b, c) + str(NewMon))
        ws['D' + str(id)] = NewMon  
        print(NewMon)
        wb.save('Data.xlsx')
        return