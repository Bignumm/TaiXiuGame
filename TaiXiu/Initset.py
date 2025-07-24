from openpyxl import Workbook
def Setup():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Acc_Fin'
    ws['A1'] = "Index"
    ws['B1'] = "Account"
    ws['C1'] = "Password"
    ws['D1'] = "Balance"
    ws['E1'] = "Count"
    ws['E2'] = 2
    ws['F1'] = "Permission"
    ws['A2'] = 1
    ws['B2'] = "admin"
    ws['C2'] = "admin"
    ws['D2'] = 10000
    ws['H1'] = 0
    ws['F2'] = 0
    wb.save('Data.xlsx')